#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug  6 15:03:36 2024

@author: ijeong-yeon
"""

"""
Basic Libraries
"""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from PIL import Image
from pandas.plotting import table
from IPython.display import clear_output
import warnings
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import azobe_stats as bs

warnings.filterwarnings("ignore")
plt.rcParams['figure.dpi'] = 500

bodhi_blue = (0.0745, 0.220, 0.396)
bodhi_grey = (0.247, 0.29, 0.322)
bodhi_primary_1 = (0.239, 0.38, 0.553)
bodhi_secondary = (0.133, 0.098, 0.42)
bodhi_tertiary = (0.047, 0.396, 0.298)
bodhi_complement = (0.604, 0.396, 0.071)

""" for breakdown or statistics 
empty_df = pd.DataFrame()
file_path = 'Visuals/Tables/stats_sheet.xlsx'   
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    empty_df.to_excel(writer, sheet_name='basic', index=False)
"""

"""
Basic visualisation setting
"""
def table_breakdown(df, dis_cols, var, sheet_name, var_name, file_path, var_order):
    dfs = {}
    book = load_workbook(file_path)
    if var_order is not None:
        df[var] = df[var].astype('category')
        df[var] = df[var].cat.set_categories(var_order, ordered=True)
    
    melted = df.melt(id_vars=dis_cols, value_vars=var, var_name=' ', value_name='category_value')
    for col, i in zip(dis_cols, range(len(dis_cols))):
        count_df = melted.groupby(['category_value', col]).size().unstack(fill_value=0)
        percent_df = round(count_df.div(count_df.sum(axis=0), axis=1) * 100, 2)
        f_df = pd.concat([count_df, percent_df.add_suffix('(%)')], axis=1)
        dfs[f'final_df{i}'] = f_df.transpose()
    final_df = pd.concat(dfs, axis=0)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        final_df.to_excel(writer, sheet_name=sheet_name, index=True, header=True)
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    ws.insert_rows(1)
    ws['B1'] = var_name
    ws['B1'].font = Font(bold=True)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
    
def count_df(df, column_name, file_path, custom_order=None, index_name='Age', region = "None"):
    count = df[column_name].value_counts()
    
    if custom_order is not None:
        count = count.loc[custom_order]
    
    count_df = pd.DataFrame({'Count': count})

    count_df['Percentage'] = round(count_df['Count'] / count_df['Count'].sum() * 100, 1)

    count_df.index.name = index_name
    if region == "None":
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            count_df.to_excel(writer, sheet_name=column_name, index=True, header=True)
        wb = load_workbook(file_path)
        ws = wb[column_name]
        ws.insert_rows(1)
        ws['A1'] = ''.join([char.upper() if char.isalpha() else char for char in column_name])
        ws['A1'].font = Font(bold=True)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    else: 
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            count_df.to_excel(writer, sheet_name=f'{column_name}_{region}', index=True, header=True)
        wb = load_workbook(file_path)
        ws = wb[column_name]
        ws.insert_rows(1)
        ws['A1'] = ''.join([char.upper() if char.isalpha() else char for char in column_name])
        ws['A1'].font = Font(bold=True)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path) 
    return count_df

def multi_table(df, columns, categories, file_path, change = None, column_labels=None, index_name = ' ', region = "None"):
    table = pd.DataFrame(index=categories)

    for col in columns:
        table[col] = df[col].value_counts().reindex(categories, fill_value=0)
    
    if column_labels is not None:
        table.columns = column_labels

    if change is not None:
        table.index = change[:len(table)]
        
    if region == "None":
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            table.to_excel(writer, sheet_name=index_name, index=True, header=True)
        wb = load_workbook(file_path)
        ws = wb[index_name]
        ws.insert_rows(1)
        ws['A1'] = ''.join([char.upper() if char.isalpha() else char for char in index_name])
        ws['A1'].font = Font(bold=True)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    else: 
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            table.to_excel(writer, sheet_name=f'{index_name}_{region}', index=True, header=True)
        wb = load_workbook(file_path)
        ws = wb[index_name]
        ws.insert_rows(1)
        ws['A1'] = ''.join([char.upper() if char.isalpha() else char for char in index_name])
        ws['A1'].font = Font(bold=True)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path) 
  
    return table

def plot_stacked_bar(df, title, output_file, xlabel = ' ', ylabel ='Count', legend_title = 'Response', rotation=45, figsize=(12, 8)):

    state_count = df.value_counts()
    
    if len(state_count) == 2:
        color_palette = [bodhi_primary_1, bodhi_complement]
    elif len(state_count) == 3:
        color_palette = [bodhi_primary_1, bodhi_complement, bodhi_tertiary]
    elif len(state_count) == 4:
        color_palette = [bodhi_primary_1, bodhi_complement, bodhi_tertiary, bodhi_blue]
    elif len(state_count) == 5:
        color_palette = [bodhi_primary_1, bodhi_complement, bodhi_tertiary, bodhi_blue, bodhi_grey]
    elif len(state_count) == 6:
        color_palette = [bodhi_primary_1, bodhi_complement, bodhi_tertiary, bodhi_blue, bodhi_grey, bodhi_secondary]
    else:
        color_palette = bodhi_blue
    
    ax = df.T.plot(kind='bar', stacked=True, figsize=figsize, color=color_palette)

    plt.title(title, fontsize=14)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.xticks(rotation=rotation, ha='center', fontsize = 14)
    
    for container in ax.containers:
        for bar in container:
            width = bar.get_width()
            height = bar.get_height()
            if height > 0:
                x = bar.get_x() + width / 2
                y = bar.get_y() + height / 2
                ax.annotate(f'{int(height)}', xy=(x, y), xytext=(0, 0), textcoords='offset points', ha='center', va='center', fontsize=11, color='black')
                
    plt.legend(title=legend_title, bbox_to_anchor=(1.05, 1), loc='upper left')

    plt.savefig(output_file, bbox_inches='tight', dpi=800)
    plt.show()

def table_with_replace(df, columns, column_labels, replace_map = {0: 'No', 1: 'Yes'}, categories=['Yes', 'No']):
    df_copy = df.copy()
    df_copy.replace(replace_map, inplace=True)
    tables = {}

    for col in columns:
        value_counts = df_copy[col].value_counts().reindex(categories, fill_value=0)
        tables[col] = pd.DataFrame(value_counts)
    
    table = pd.concat(tables, axis=1)
    table.columns = table.columns.droplevel(1)
    table.columns = column_labels
    
    return table

def plot_bar(df, col, title, output_file, xlabel=' ', ylabel='Percentage', figsize=(12, 8), rotation=45, fontsize=12):
    
    plt.figure(figsize=figsize)

    state_count = df.value_counts()
    
    if len(state_count) <= 6:
        palette = [bodhi_complement, bodhi_blue, bodhi_tertiary, bodhi_primary_1, bodhi_grey, bodhi_secondary]
    else:
        palette = bodhi_blue
    df2 = df['Count']
    df = df[col]
    bars = df.plot(kind='bar', color=palette)

    for bar, value in zip(bars.patches, df2.values):   
        
        percentage = (value / df2.values.sum()) * 100
        plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), 
             f'{value} ({percentage:.1f}%)',
             ha='center', va='bottom', fontsize=fontsize+2)
    
        

    plt.title(title, fontsize=fontsize + 4)
    plt.xlabel(xlabel, fontsize = fontsize)
    plt.ylabel(ylabel, fontsize = fontsize)
    plt.xticks(rotation=rotation, fontsize=fontsize+4)
    plt.ylim(0, 100)
    

    plt.savefig(output_file, bbox_inches='tight', dpi=800)
    plt.show()

def top_plot(df, title, output_file, number = 2, rotation=0, fontsize = 14):
    yes_row = df.loc['Yes']
    top_columns = yes_row.sort_values(ascending=False)[0:number]

    if number == 2:
        color_palette = [bodhi_primary_1, bodhi_complement]
    elif number == 3:
        color_palette = [bodhi_primary_1, bodhi_complement, bodhi_tertiary]
    elif number == 4:
        color_palette = [bodhi_primary_1, bodhi_complement, bodhi_tertiary, bodhi_blue]
    elif number == 5:
        color_palette = [bodhi_primary_1, bodhi_complement, bodhi_tertiary, bodhi_blue, bodhi_grey]
    elif number == 6:
        color_palette = [bodhi_primary_1, bodhi_complement, bodhi_tertiary, bodhi_blue, bodhi_grey, bodhi_secondary]
    else:
        color_palette = bodhi_blue

    plt.figure(figsize=(10, 6))
    bars = plt.bar(top_columns.index, top_columns.values, color=color_palette)

    for bar, value in zip(bars.patches, top_columns.values):

        
        percentage = (value / top_columns.values.sum()) * 100
        plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), 
             f'{value} ({percentage:.1f}%)',
             ha='center', va='bottom', fontsize = fontsize)

    plt.xlabel(' ')
    plt.ylabel('Values')
    plt.title(title, fontsize = fontsize+2)
    plt.xticks(rotation=rotation, fontsize = fontsize)
    
    plt.legend(fontsize=fontsize, loc='best')
    
    plt.savefig(output_file, bbox_inches='tight', dpi=800)
    plt.show()
 
"""
Functions for each indicator
"""
 
def leao111(cay_df, file_path, breakdown = None):
    def leao111_score(row):
        score_mapping = {
        'Strongly agree': 3,
        'Agree': 2,
        'Disagree': 1,
        'Strongly disagree': 0
        }
        score = 0
        columns = ['21-1', '21-2', '21-3', '21-4', '21-5']
        for col in columns:
            score += score_mapping.get(row[col], 0)
        return score / len(columns)

    cay_df['average_score'] = cay_df.apply(leao111_score, axis=1)
    cay_df['leao1.1.1'] = cay_df['average_score'].apply(lambda x: 'Empowered' if x >= 2 else 'Not empowered')
    cay_df.drop(columns=['average_score'], inplace = True)
    
    if breakdown == None:
        bs.chi2(cay_df, 'leao1.1.1', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)

    order= ['Empowered','Not empowered']
    xlabel = ' '
    if breakdown == None:
        leao111_df = count_df(cay_df, 'leao1.1.1', file_path, custom_order=order, index_name='LEAO1.1.1')
        title = 'LEAO1.1.1: % of young people who demonstrate empowerment'
        plot_bar(leao111_df, 'Percentage', title = title, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_LEAO1.1.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'leao1.1.1'
        var_name = "LEAO1.1.1: % of young people who demonstrate empowerment"
        leao111_df = table_breakdown(cay_df, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
            
def leao211(cay_df, file_path, breakdown = None):
    def leao211_score(row):
        score_mapping = {
        'Very supportive': 3,
        'Extremely supportive': 2,
        'Somewhat supportive': 1,
        'Not supportive at all': 0
        }
        if row.isnull().all():
            return np.nan 
        score = 0
        columns = ['22-1', '22-2', '22-3', '22-4']
        non_na_count = 0  
    
        for col in columns:
            if pd.notna(row[col]):
                score += score_mapping.get(row[col], 0)
                non_na_count += 1
        if non_na_count == 0:
            return np.nan
        if len(columns) - non_na_count <= 2:
            return score / len(columns)
        else:
            return score / non_na_count
    cay_df['average_score'] = cay_df.apply(leao211_score, axis=1)
    cay_df['leao2.1.1'] = cay_df['average_score'].apply(lambda x: 'Supportive' if pd.notna(x) and x >= 3 else 'Not supportive' if pd.notna(x) else np.nan)
    cay_df.drop(columns=['average_score'], inplace=True)
    
    if breakdown == None:
        bs.chi2(cay_df, 'leao2.1.1', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
    
    order= ['Supportive','Not supportive']
    xlabel = ' '

    if breakdown == None:
        leao211_df = count_df(cay_df, 'leao2.1.1', file_path, custom_order=order, index_name='LEAO2.1.1')
        title = 'LEAO2.1.1: % of young people who report that their parents/caregivers support \nand publicly defend their engagement in decision-making processes at different levels'
        plot_bar(leao211_df, 'Percentage', title = title, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_LEAO2.1.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'leao2.1.1'
        var_name = "LEAO2.1.1: % of young people who report that their parents/caregivers support and publicly defend their engagement in decision-making processes at different levels"
        leao211_df = table_breakdown(cay_df, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def srhi111(cay_df, file_path, breakdown = None):
    srhi111 = cay_df[(cay_df['4'] == 'Female') & (cay_df['2'] >= 15) & (cay_df['2'] <= 24) & (cay_df['23'] == 'Yes') & (cay_df['24'] == 'Yes')]
    condition = srhi111[['25-1', '25-2', '25-3', '25-4', '25-5', '25-o']].eq('Yes').any(axis=1)
    srhi111['srhi1.1.1'] = np.where(condition, 'Use', 'Do not use')
    
    order= ['Use', 'Do not use']
    xlabel = ' '
    if breakdown == None:
        srhi111_df = count_df(srhi111, 'srhi1.1.1', file_path, custom_order=order, index_name='SRHI1.1.1')
        title = 'SRHI1.1.1: % of sexually active young women aged 15-24 \nwho are currently using a modern method of contraception'
        plot_bar(srhi111_df, 'Percentage', title = title, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHI1.1.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srhi1.1.1'
        var_name = "SRHI1.1.1: % of sexually active young women aged 15-24 who are currently using a modern method of contraception"
        srhi111_df = table_breakdown(srhi111, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def srhi112(cay_df, file_path, breakdown = None):
    srhi112 = cay_df[(cay_df['4'] == 'Female') & (cay_df['2'] <= 20) & (cay_df['26'] == 'Yes')]

    condition = srhi112[['28-B-1', '28-B-2', '28-B-3']].eq('No').any(axis=1)

    srhi112['srhi1.1.2'] = np.where(condition, 'Unintended', 'Intended')
    
    order= ['Unintended','Intended']
    xlabel = ' '
    
    if breakdown == None:
        srhi112_df = count_df(srhi112, 'srhi1.1.2', file_path, custom_order=order, index_name='SRHI1.1.2')
        title = 'SRHI1.1.2: % of recent births that are reported as unintended to mothers under 20'
        plot_bar(srhi112_df, 'Percentage', title = title, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHI1.1.2.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srhi1.1.2'
        var_name = "SRHI1.1.2: % of recent births that are reported as unintended to mothers under 20"
        srhi112_df = table_breakdown(srhi112, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
 
def srhi311(cay_df, file_path, breakdown = None):
    srhi311 = cay_df[(cay_df['4'] == 'Female') & (cay_df['2'] >= 15) &(cay_df['2'] <= 24)]
    srhi311['14'] = pd.to_numeric(srhi311['14'], errors='coerce')

    conditions = [
    (srhi311['14'] <= 15),
    (srhi311['14'] > 15) & (srhi311['14'] <= 18)]
    choices = ['Before 15', 'Before 18']

    srhi311['srhi3.1.1'] = np.select(conditions, choices, default='After 18')
    
    order= ['Before 15','Before 18','After 18']
    xlabel = ' '
    if breakdown == None:
        srhi311_df = count_df(srhi311, 'srhi3.1.1', file_path, custom_order=order, index_name='SRHI3.1.1')
        title = 'SRHI3.1.1: % of girls and women aged 15-24 years who were married'
        plot_bar(srhi311_df, 'Percentage', title = title, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHI3.1.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srhi3.1.1'
        var_name = "SRHI3.1.1: % of girls and women aged 15-24 years who were married"
        srhi311_df = table_breakdown(srhi311, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def srhi114(cay_df, file_path, breakdown = None):

    srhi114 = cay_df[(cay_df['4'] == 'Female') & (cay_df['2'] >= 20) &(cay_df['2'] <= 24)]
    srhi114['14'] = pd.to_numeric(srhi114['14'], errors='coerce')

    conditions = [
    (srhi114['14'] <= 15),
    (srhi114['14'] > 15) & (srhi114['14'] <= 18)]
    choices = ['Before 15', 'Before 18']

    srhi114['srhi1.1.4'] = np.select(conditions, choices, default='After 18')
    
    order= ['Before 15','Before 18','After 18']
    xlabel = ' '
    if breakdown == None:
        srhi114_df = count_df(srhi114, 'srhi1.1.4', file_path, custom_order=None, index_name='SRHI1.1.4')
        title = 'SRHSA1.1.4: % of women aged 20-24 years who were married or in a union a) before the age of 15, and b) before the age of 18'
        plot_bar(srhi114_df, 'Percentage', title = title, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHI1.1.4.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srhi1.1.4'
        var_name = "SRHSA1.1.4: % of women aged 20-24 years who were married or in a union a) before the age of 15, and b) before the age of 18"
        srhi114_df = table_breakdown(srhi114, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def ecdi323(cay_df, file_path, breakdown = None):
    ecdi323 = cay_df[(cay_df['4'] == 'Female')& (cay_df['29'] == 'Yes')]
    ecdi323['30'] = pd.to_numeric(ecdi323['31'], errors='coerce')
    ecdi323['31'] = pd.to_numeric(ecdi323['31'], errors='coerce')

    condition = (ecdi323['30'] >= 4) & (ecdi323['31'] >= 1)

    ecdi323['ecdi3.2.3'] = np.where(condition, 'At least 4 contact', 'Less than 4 contacts')
    
    order= ['At least 4 contact','Less than 4 contacts']
    xlabel = " "
    if breakdown == None:
        ecdi323_df = count_df(ecdi323, 'ecdi3.2.3', file_path, custom_order=order, index_name='ECDI3.2.3')
        title = 'ECDI3.2.3: % of adolescents and women with a birth in the last two years who had at least \nfour antenatal contacts, including at least one in the last trimester, during the last pregnancy'
        plot_bar(ecdi323_df, 'Percentage', title = title, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_ECDI3.2.3.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:   
        var = 'ecdi3.2.3'
        var_name = "ECDI3.2.3: % of adolescents and women with a birth in the last two years who had at least four antenatal contacts, including at least one in the last trimester, during the last pregnancy"
        ecdi323_df = table_breakdown(ecdi323, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def ecdi324(cay_df, file_path, breakdown = None):
    ecdi324 = cay_df[(cay_df['4'] == 'Female')& (cay_df['29'] == 'Yes')]

    condition = ecdi324['32'] == 'Yes'

    ecdi324['ecdi3.2.4'] = np.where(condition, 'pass', 'fail')
    
    bs.chi2(ecdi324, 'ecdi3.2.4', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
    
    order = ['pass'] # Need to put fail in the future
    xlabel = ' '
    if breakdown == None:
        ecdi324_df = count_df(ecdi324, 'ecdi3.2.4', file_path, custom_order=order, index_name='ECDI3.2.4')
        title = 'ECDI3.2.4: % of adolescents and women with a live birth (in the last 2 years)\nwho were attended by skilled health personnel during delivery'
        plot_bar(ecdi324_df, 'Percentage', title = title, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_ECDI3.2.4.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'ecdi3.2.4'
        var_name = "ECDI3.2.4: % of adolescents and women with a live birth (in the last 2 years) who were attended by skilled health personnel during delivery"
        ecdi324_df = table_breakdown(ecdi324, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
         
def srho111_separated(cay_df, file_path, breakdown = None):
    srho111 = cay_df[(cay_df['2'] >= 13) &(cay_df['2'] <= 24)]

    srho111_1 = srho111[(srho111['33'] == 'Yes')]
    columns_to_check = ['34-1', '34-2', '34-3', '34-4', '34-5', '34-6', '34-7', '34-8', '34-o']
    srho111_1['srho1.1.1_1'] = srho111_1[columns_to_check].sum(axis=1).apply(lambda x: 'Pass' if x >= 3 else 'Fail')

    srho111_2 = srho111[(srho111['35'] == 'Yes')]
    columns_to_check = ['36-1', '36-2', '36-3', '36-4', '36-5', '36-6', '36-7', '36-8', '36-o']
    srho111_2['srho1.1.1_2'] = srho111_2[columns_to_check].sum(axis=1).apply(lambda x: 'Pass' if x >= 3 else 'Fail')

    srho111_3 = srho111[(srho111['37'] == 'Yes')]
    columns_to_check = ['38-1','38-2','38-3','38-4','38-5','38-6','38-7','38-8','38-9','38-o']
    srho111_3['srho1.1.1_3'] = srho111_3[columns_to_check].sum(axis=1).apply(lambda x: 'Pass' if x >= 3 else 'Fail')

    srho111_4 = srho111[(srho111['43'] == 'Yes')]
    srho111_4['srho1.1.1_4'] = srho111_4['45'].apply(lambda x: 'Pass' if x =='Yes' else 'Fail')

    srho111_5 = srho111.copy(deep=True)
    columns_to_check = ['50', '51', '52', '53', '54', '55']
    srho111_5['srho1.1.1_5'] = srho111_5[columns_to_check].apply(lambda row: 'Pass' if sum(row == 'True') >= 3 else 'Fail', axis=1)

    srho111_6 = srho111.copy(deep=True)
    srho111_6['srho1.1.1_6'] = srho111_6['56'].apply(lambda x: 'Pass' if x =='False' else 'Fail')

    order= ['Pass','Fail']
    xlabel = ' '
    if breakdown == None:
        srho111_1_df = count_df(srho111_1, 'srho1.1.1_1', file_path, custom_order=order, index_name='SRHO1.1.1-1')
        srho111_2_df = count_df(srho111_2, 'srho1.1.1_2', file_path, custom_order=order, index_name='SRHO1.1.1-2')
        srho111_3_df = count_df(srho111_3, 'srho1.1.1_3', file_path, custom_order=order, index_name='SRHO1.1.1-3')
        srho111_4_df = count_df(srho111_4, 'srho1.1.1_4', file_path, custom_order=order, index_name='SRHO1.1.1-4')
        srho111_5_df = count_df(srho111_5, 'srho1.1.1_5', file_path, custom_order=order, index_name='SRHO1.1.1-5')
        srho111_6_df = count_df(srho111_6, 'srho1.1.1_6', file_path, custom_order=order, index_name='SRHO1.1.1-6')

        title1 = 'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Physical changes during puberty for boys>'
        title2 = 'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Physical changes during puberty for girls>'
        title3 = 'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Knowledge of contraceptive methods>'
        title4 = 'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Knowledge of preventing HIV transmission>'
        title5 = 'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Reproductive health>'
        title6 = 'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<People’s private behaviour>'

        plot_bar(srho111_1_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.1.1-1.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_2_df, 'Percentage', title = title2, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.1.1-2.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_3_df, 'Percentage', title = title3, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.1.1-3.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_4_df, 'Percentage', title = title4, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.1.1-4.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_5_df, 'Percentage', title = title5, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.1.1-5.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_6_df, 'Percentage', title = title6, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.1.1-6.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        srho111_1_df = count_df(srho111_1, 'srho1.1.1_1', file_path, custom_order=order, index_name=f'SRHO1.1.1-1_{breakdown}')
        srho111_2_df = count_df(srho111_2, 'srho1.1.1_2', file_path, custom_order=order, index_name=f'SRHO1.1.1-2_{breakdown}')
        srho111_3_df = count_df(srho111_3, 'srho1.1.1_3', file_path, custom_order=order, index_name=f'SRHO1.1.1-3_{breakdown}')
        srho111_4_df = count_df(srho111_4, 'srho1.1.1_4', file_path, custom_order=order, index_name=f'SRHO1.1.1-4_{breakdown}')
        srho111_5_df = count_df(srho111_5, 'srho1.1.1_5', file_path, custom_order=order, index_name=f'SRHO1.1.1-5_{breakdown}')
        srho111_6_df = count_df(srho111_6, 'srho1.1.1_6', file_path, custom_order=order, index_name=f'SRHO1.1.1-6_{breakdown}')

        title1 = f'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Physical changes during puberty for boys>'
        title2 = f'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Physical changes during puberty for girls>'
        title3 = f'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Knowledge of contraceptive methods>'
        title4 = f'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Knowledge of preventing HIV transmission>'
        title5 = f'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<Reproductive health>'
        title6 = f'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics \n<People’s private behaviour>'
        plot_bar(srho111_1_df, 'Percentage', title = title1, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_SRHO1.1.1-1.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_2_df, 'Percentage', title = title2, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_SRHO1.1.1-2.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_3_df, 'Percentage', title = title3, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_SRHO1.1.1-3.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_4_df, 'Percentage', title = title4, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_SRHO1.1.1-4.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_5_df, 'Percentage', title = title5, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_SRHO1.1.1-5.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(srho111_6_df, 'Percentage', title = title6, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_SRHO1.1.1-6.png', figsize=(18, 10), rotation=0, fontsize = 14)

    clear_output()

def srho111(cay_df, file_path, breakdown = None):

    srho111 = cay_df[(cay_df['2'] >= 13) &(cay_df['2'] <= 24)]

    srho111_1 = srho111[(srho111['33'] == 'Yes')]
    columns_to_check = ['34-1', '34-2', '34-3', '34-4', '34-5', '34-6', '34-7', '34-8', '34-o']
    srho111_1['srho1.1.1_1'] = srho111_1[columns_to_check].sum(axis=1).apply(lambda x: 'Pass' if x >= 3 else 'Fail')

    srho111_2 = srho111[(srho111['35'] == 'Yes')]
    columns_to_check = ['36-1', '36-2', '36-3', '36-4', '36-5', '36-6', '36-7', '36-8', '36-o']
    srho111_2['srho1.1.1_2'] = srho111_2[columns_to_check].sum(axis=1).apply(lambda x: 'Pass' if x >= 3 else 'Fail')

    srho111_3 = srho111[(srho111['37'] == 'Yes')]
    columns_to_check = ['38-1','38-2','38-3','38-4','38-5','38-6','38-7','38-8','38-9','38-o']
    srho111_3['srho1.1.1_3'] = srho111_3[columns_to_check].sum(axis=1).apply(lambda x: 'Pass' if x >= 3 else 'Fail')

    srho111_4 = srho111[(srho111['43'] == 'Yes')]
    srho111_4['srho1.1.1_4'] = srho111_4['45'].apply(lambda x: 'Pass' if x =='Yes' else 'Fail')

    srho111_5 = srho111.copy(deep=True)
    columns_to_check = ['50', '51', '52', '53', '54', '55']
    srho111_5['srho1.1.1_5'] = srho111_5[columns_to_check].apply(lambda row: 'Pass' if sum(row == 'True') >= 3 else 'Fail', axis=1)

    srho111_6 = srho111.copy(deep=True)
    srho111_6['srho1.1.1_6'] = srho111_6['56'].apply(lambda x: 'Pass' if x =='False' else 'Fail')

    srho111['srho1.1.1_dc'] = 0
    result = srho111_1[srho111_1['srho1.1.1_1'] == 'Pass']
    result = result.merge(srho111_2[srho111_2['srho1.1.1_2'] == 'Pass'], how='inner')
    result = result.merge(srho111_3[srho111_3['srho1.1.1_3'] == 'Pass'], how='inner')
    result = result.merge(srho111_4[srho111_4['srho1.1.1_4'] == 'Pass'], how='inner')
    result = result.merge(srho111_5[srho111_5['srho1.1.1_5'] == 'Pass'], how='inner')
    resultss = result.merge(srho111_6[srho111_6['srho1.1.1_6'] == 'Pass'], how='inner')
    
    dataframes = {
    'srho1.1.1_1': srho111_1,
    'srho1.1.1_2': srho111_2,
    'srho1.1.1_3': srho111_3,
    'srho1.1.1_4': srho111_4,
    'srho1.1.1_5': srho111_5,
    'srho1.1.1_6': srho111_6
    }

    indices_pass = {}
    for col, df in dataframes.items():
        indices_pass[col] = set(df[df[col] == 'Pass'].index)

    common_indices = set.intersection(*indices_pass.values())
    common_indices = list(common_indices)

    srho111.loc[common_indices, 'srho1.1.1_dc'] = 1
    
    if breakdown == None:
        bs.chi2(srho111, 'srho1.1.1_dc', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)

    len_result = len(srho111)
    len_pass = len(resultss)
    len_fail = len_result - len_pass
    count_pass = len_pass
    count_fail = len_fail
    percent_pass = (count_pass / len_result) * 100
    percent_fail = (count_fail / len_result) * 100
    result_srho111_1 = pd.DataFrame({
        'Count': [count_pass, count_fail],
        'Percent': [percent_pass, percent_fail]
    }, index=['pass', 'fail'])
    xlabel = ' '
    if breakdown == None:
        print(f'{percent_pass:.2f} percent of people meet the criteria for indicator SRHO1.1.1')
        title1 = 'SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics'
        plot_bar(result_srho111_1, 'Percent', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.1.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srho1.1.1_dc'
        var_name = "SRHO1.1.1: % of CAY with correct knowledge about SRHR core topics"
        srho111_df = table_breakdown(srho111, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = None)
  
def srho112(cay_df, file_path, breakdown = None):
    srho112 = cay_df[(cay_df['2'] >= 10) &(cay_df['2'] <= 24)]

    def calculate_score(row):
        score = 0
        score += sum([1 for col in ['50', '52', '53','54', '55'] if row[col] == 'True'])
        score += sum([1 for col in ['51', '56'] if row[col] == 'False'])
        return score
    
    def label_score(score):
        if score <= 2:
            return 'No Knowledge'
        elif score == 3:
            return 'Basic'
        elif score in [4, 5]:
            return 'Moderate'
        else:
            return 'High'

    srho112['score'] = srho112.apply(calculate_score, axis=1)
    srho112['srho.1.1.2'] = ""
    srho112['srho.1.1.2'] = srho112['score'].apply(label_score)
    srho112.drop(columns=['score'], inplace=True)
    
    if breakdown == None:
        bs.chi2(srho112, 'srho.1.1.2', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
    
    order= ['No Knowledge','Basic','Moderate','High']
    xlabel = ' '
    if breakdown == None:
        srho112_df = count_df(srho112, 'srho.1.1.2', file_path, custom_order=order, index_name='SRHO1.1.2')
        title1 = 'SRHO1.1.2 % of CAY who have at least moderate levels of knowledge about menstrual healths'
        plot_bar(srho112_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.1.2.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srho.1.1.2'
        var_name = "SRHO1.1.2 % of CAY who have at least moderate levels of knowledge about menstrual healths"
        srho112_df = table_breakdown(srho112, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
 
def srho113(cay_df, file_path, breakdown = None):
    srho113 = cay_df[(cay_df['2'] >= 15) &(cay_df['2'] <= 24)]

    def calculate_score(row):
        score = 0
        score += sum([1 for col in ['39', '42', '44'] if row[col] == 'Yes'])
        score += sum([1 for col in ['40', '41'] if row[col] == 'No'])
        return score
    
    def label_score(score):
        if score == 5:
            return 'Pass'
        else:
            return 'Fail'

    srho113['score'] = srho113.apply(calculate_score, axis=1)
    srho113['srho.1.1.3'] = ""
    srho113['srho.1.1.3'] = srho113['score'].apply(label_score)
    srho113.drop(columns=['score'], inplace=True)
    
    if breakdown == None:
        bs.chi2(srho113, 'srho.1.1.3', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
    
    order= ['Pass','Fail']
    xlabel = ' '
    if breakdown == None:
        srho113_df = count_df(srho113, 'srho.1.1.3', file_path, custom_order=order, index_name='SRHO1.1.3')
        title1 = 'SRHO1.1.3: % of young women and men aged 15- 24 who correctly identify ways of preventing \nthe sexual transmission of HIV and who reject major misconceptions about HIV transmission'
        plot_bar(srho113_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.1.3.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srho.1.1.3'
        var_name = "SRHO1.1.3 % of young women and men aged 15- 24 who correctly identify ways of preventing the sexual transmission of HIV and who reject major misconceptions about HIV transmission"
        srho113_df = table_breakdown(srho113, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
 
def srho121(cay_df, file_path, breakdown = None):
    srho121 = cay_df[(cay_df['2'] >= 15) &(cay_df['2'] <= 24)]

    def calculate_score(row):
        score = 0
        score += sum([1 for col in ['57', '58-1', '58-2','59-1','59-2'] if row[col] == 'Yes'])
        return score
    
    def label_score(score):
        if score == 3:
            return 'Confident'
        else:
            return 'Not confident'

    srho121['score'] = srho121.apply(calculate_score, axis=1)
    srho121['srho.1.2.1'] = ""
    srho121['srho.1.2.1'] = srho121['score'].apply(label_score)
    srho121.drop(columns=['score'], inplace=True)
    
    if breakdown == None:
        bs.chi2(srho121, 'srho.1.2.1', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
    
    
    order= ['Confident','Not confident']
    xlabel = ' '
    if breakdown == None:
        srho121_df = count_df(srho121, 'srho.1.2.1', file_path, custom_order=order, index_name='SRHO1.2.1')
        title1 = 'SRHO1.2.1: % (and #) of adolescents and youth aged 15 - 24 who feel able to \nmake informed decisions about their sexual and reproductive health'
        plot_bar(srho121_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.2.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srho.1.2.1'
        var_name = "SRHO1.2.1: % (and #) of adolescents and youth aged 15 - 24 who feel able to make informed decisions about their sexual and reproductive health"
        srho121_df = table_breakdown(srho121, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
         
def srho141(cay_df, file_path, breakdown = None):
    srho141 = cay_df[(cay_df['2'] >= 15) &(cay_df['2'] <= 24)]

    def assign_pass_fail(row):
        if row['46'] == 'Yes':
            return 'Can get condoms'
        else:
            return 'Cannot get condoms'
    srho141['srho.1.4.1'] = ""
    srho141['srho.1.4.1'] = srho141.apply(assign_pass_fail, axis=1)
    
    if breakdown == None:
        bs.chi2(srho141, 'srho.1.4.1', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
    
    order= ['Can get condoms','Cannot get condoms']
    xlabel = ' '
    if breakdown == None:
        srho141_df = count_df(srho141, 'srho.1.4.1', file_path, custom_order=order, index_name='SRHO1.4.1')
        title1 = 'SRHO1.4.1: % of adolescents and youth who say that they could get condoms'
        plot_bar(srho141_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.4.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srho.1.4.1'
        var_name = "SRHO1.4.1: % of adolescents and youth who say that they could get condoms"
        srho141_df = table_breakdown(srho141, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
         
def srho144(cay_df, file_path, breakdown = None):
    srho144 = cay_df[(cay_df['2'] >= 15) &(cay_df['2'] <= 49)]
    def assign_pass_fail(row):
        if (row['48'] == 'Yes') & (row['49'] == 'Yes'):
            return 'Yes'
        else:
            return 'No'

    srho144['srho.1.4.4'] = srho144.apply(assign_pass_fail, axis=1)
    
    bs.chi2(srho144, 'srho.1.4.4', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
    
    order= ['Yes','No']
    xlabel = ' '
    if breakdown == None:
        srho144_df = count_df(srho144, 'srho.1.4.4', file_path, custom_order=order, index_name='SRHO1.4.4')
        title1 = 'SRHO1.4.4: % of adolescents and youth who were tested for HIV and \nreceived their results during the reporting period'
        plot_bar(srho144_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO1.4.4.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srho.1.4.4'
        var_name = "SRHO1.4.4: % of adolescents and youth who were tested for HIV and received their results during the reporting period"
        srho144_df = table_breakdown(srho144, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def srho621(cay_df, file_path, breakdown = None):
    def srho621_score(row):
        score_mapping_1 = {
        'Agree': 1,
        'Disagree': 0,
        'Unsure': 0,
        'Neither agree nor disagree': 0, 'Prefer not to answer': 0
        }
        score_mapping_2 = {
        'Agree': 0.5,
        'Disagree': 0,
        'Unsure': 0,
        'Neither agree nor disagree': 0, 'Prefer not to answer': 0
        }
        score = 0
        columns1 = ['60','61','62','63','65','67','68']
        columns2 = ['64', '66','69-1','69-2']
        for col in columns1:
            score += score_mapping_1.get(row[col], 0)
        for col in columns2:
            score += score_mapping_2.get(row[col], 0)
        return score

    def label_score(score):
        if score <= 3:
            return 'Not responsive'
        elif 3 < score <= 6.5:
            return 'Partially responsive'
        else:
            return 'Highly responsive'
    srho621 = cay_df[(cay_df['2'] >= 15) &(cay_df['2'] <= 24)]
    srho621['score'] = srho621.apply(srho621_score, axis=1)
    srho621['srho6.2.1'] = ""
    srho621['srho6.2.1'] = srho621['score'].apply(label_score)
    srho621.drop(columns=['score'], inplace = True)
    
    if breakdown == None:
        bs.chi2(srho621, 'srho6.2.1', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)

    order= ['Not responsive','Partially responsive','Highly responsive']
    xlabel = ' '
    if breakdown == None:
        srho621_df = count_df(srho621, 'srho6.2.1', file_path, custom_order=order, index_name='SRHO6.2.1')
        title1 = 'SRHO6.2.1: % (and #) of adolescents and youth aged 15 – 24 \nwho consider services to be adolescent and gender-responsive at the time asked'
        plot_bar(srho621_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SRHO6.2.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'srho6.2.1'
        var_name = "SRHO6.2.1: % (and #) of adolescents and youth aged 15 – 24 who consider services to be adolescent and gender-responsive at the time asked'"
        srho621_df = table_breakdown(srho621, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
         
def proi111(cay_df, file_path, breakdown = None):
    proi111 = cay_df[(cay_df['2'] <= 18)]

    def assign_label(row):
        if any(row[col] == 'Yes' for col in ['70','71','72','73']):
            return 'Exist'
        else:
            return 'None'
    proi111['proi1.1.1'] = ""
    proi111['proi1.1.1'] = proi111.apply(assign_label, axis=1)
    
    if breakdown == None:
        bs.chi2(proi111, 'proi1.1.1', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
        
        
    order= ['Exist','None']
    xlabel = ""
    if breakdown == None:
        proi111_df = count_df(proi111, 'proi1.1.1', file_path, custom_order=order, index_name='PROI1.1.1')
        title1 = 'PROI1.1.1: % of children who report incidents of \nviolence, abuse, exploitation and neglect within 12 months of program implementation'
        plot_bar(proi111_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROI1.1.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'proi1.1.1'
        var_name = "PROI1.1.1: % of children who report incidents of violence, abuse, exploitation and neglect within 12 months of program implementation"
        proi111_df = table_breakdown(proi111, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
  
def proo111(cay_df, file_path, breakdown = None):
    proo111 = cay_df[(cay_df['2'] >= 10) & (cay_df['2'] <= 24)]

    def proo111_score_a(row):
        score = 0
        score += sum([1 for col in ['74-1','74-2','74-3','74-4','74-5','74-6','74-o'] if row[col] == 1])
        return score

    def proo111_score_bc(row):
        score_mapping_1 = {
        'Yes': 1,
        'No': 0,
        'Unsure': 0,
        'Prefer not to answer': 0
        }

        score = 0
        columns1 = ['75','77','79', '81','83','85','87']
        for col in columns1:
            score += score_mapping_1.get(row[col], 0)
        if any(row[col] == 1 for col in ['84-1','84-2','84-3','84-4','84-5','84-6','84-7','84-o']):
            score += 1
        return score

    def label_score(score):
        if score < 3:
            return 'Insufficient'
        elif score >= 3:
            return 'Sufficient'

    def label_score2(score):
        if score < 4:
            return 'Insufficient'
        elif score >= 4:
            return 'Sufficient'
        
    proo111['scorea'] = proo111.apply(proo111_score_a, axis=1)
    proo111['proo1.1.1_a'] = proo111['scorea'].apply(label_score)
    proo111.drop(columns=['scorea'], inplace=True)

    proo111['scoreb'] = proo111.apply(proo111_score_bc, axis=1)
    proo111['proo1.1.1_bc'] = proo111['scoreb'].apply(label_score2)
    proo111.drop(columns=['scoreb'], inplace=True)
    proo111['proo1.1.1'] = proo111.apply(lambda row: 'Sufficient' if row['proo1.1.1_a'] == 'Sufficient' and row['proo1.1.1_bc'] == 'Sufficient' else 'Insufficient', axis=1)

    if breakdown == None:
        bs.chi2(proo111, 'proo1.1.1', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)

    order= ['Sufficient','Insufficient']
    xlabel = ' '
    if breakdown == None:
        proo111_df = count_df(proo111, 'proo1.1.1', file_path, custom_order=order, index_name='PROO1.1.1')
        title1 = 'PROO1.1.1 % of CAY who demonstrate knowledge of child protection risks and behaviours'
        plot_bar(proo111_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO1.1.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'proo1.1.1'
        var_name = "PROO1.1.1 % of CAY who demonstrate knowledge of child protection risks and behaviours"
        proo111_df = table_breakdown(proo111, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
         
def proo141(cay_df, file_path, breakdown = None):
    proo141 = cay_df.copy(deep=True)

    def assign_pass_fail(row):
        if (row['89'] == 'Yes') & (row['91'] == 'Yes')& (row['93'] == 'Yes')& (row['95'] == 'Yes'):
            return 'Confident'
        else:
            return 'Not confident'
    proo141['proo1.4.1'] = ""

    proo141['proo1.4.1'] = proo141.apply(assign_pass_fail, axis=1)
    categories = [1, 0]
    change = ['Yes', 'No']
    columns_labels1 = ['Family Support Unit', 'Police', 'Social welfare', 'Report to the teachers', 
                  'Report to parents', 'Chief', 'Community or religious leader','Other']
    columns_labels2 = ['Family Support Unit', 'Police', 'Social welfare', 'Report to the teachers', 
                  'Report to parents or relatives', 'Report to friends','Chief', 'Community leader','Religious leader','Other']

    #mse_table = multi_table(proo141, ['90-1','90-2','90-3','90-4','90-5','90-6','90-7','90-o'], categories = categories, change = change, column_labels = columns_labels1, index_name = 'PROO1.4.1_90')
    #mse_table2 = multi_table(proo141, ['92-1','92-2','92-3','92-4','92-5','92-6','92-7','92-8','92-9','92-o'], categories = categories, change = change, column_labels = columns_labels2, index_name = 'PROO1.4.1_92')
    #mse_table3 = multi_table(proo141, ['94-1','94-2','94-3','94-4','94-5','94-6','94-7','94-8','94-9','94-o'], categories = categories, change = change, column_labels = columns_labels2, index_name = 'PROO1.4.1_94')
   # mse_table4 = multi_table(proo141, ['96-1','96-2','96-3','96-4','96-5','96-6','96-7','96-8','96-9','96-o'], categories = categories, change = change, column_labels = columns_labels2, index_name = 'PROO1.4.1_96')

    if breakdown == None:
        bs.chi2(proo141, 'proo1.4.1', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)

    order= ['Confident','Not confident']
    xlabel = ' '
    if breakdown == None:
        proo141_df = count_df(proo141, 'proo1.4.1', file_path, custom_order=order, index_name='PROO1.4.1')
        title1 = 'PROO1.4.1: % of CAY who report that they are confident to report a protection violation to a reporting structure'
        plot_bar(proo141_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO1.4.1.png', figsize=(18, 10), rotation=0, fontsize = 14)        
    else:
        var = 'proo1.4.1'
        var_name = "PROO1.4.1: % of CAY who report that they are confident to report a protection violation to a reporting structure"
        proo141_df = table_breakdown(proo141, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
      
def proo142(cay_df, file_path, breakdown = None):
    proo142 = cay_df.copy(deep=True)

    def assign_pass_fail(row):
        if (row['97'] == 'Yes') & (row['99'] == 'Yes'):
            return 'Know'
        else:
            return 'Do not know'
    proo142['proo1.4.2'] = ""
    proo142['proo1.4.2'] = proo142.apply(assign_pass_fail, axis=1)
    
    if breakdown == None:
        bs.chi2(proo142, 'proo1.4.2', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
        
    order= ['Know','Do not know']
    xlabel = ' '
    if breakdown == None:
        proo142_df = count_df(proo142, 'proo1.4.2', file_path, custom_order=order, index_name='PROO1.4.2')
        title1 = 'PROO1.4.2: % of CAY who report understanding \nwhere to get information and services before, during and after crises'
        plot_bar(proo142_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO1.4.2.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'proo1.4.2'
        var_name = "PROO1.4.2: % of CAY who report understanding where to get information and services before, during and after crises"
        proo142_df = table_breakdown(proo142, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def iqei121(pp_df, file_path, breakdown = None):
    iqei121 = pp_df[pp_df['15']!='No']

    def calculate_score1(row):
        condition_1 = (row['17-I'] == 'Yes' and 3 <= row['20-I'] <= 5)
        condition_2 = (row['21-I'] == 'Yes' and 6 <= row['22-I'] <= 8)
        if row['17-I'] == 'Yes' and row['21-I'] == 'Yes':
            if condition_1 and condition_2:
                return 1
            else:
                return 0
        elif condition_1 or condition_2:
            return 1
        else:
            return 0

    def calculate_score2(row):
        if pd.isna(row['17-II']):
            return np.nan
    
        condition_1 = (row['17-II'] == 'Yes' and 3 <= row['20-II'] <= 5)
        condition_2 = (row['21-II'] == 'Yes' and 6 <= row['22-II'] <= 8)
    
        if row['17-II'] == 'Yes' and row['21-II'] == 'Yes':
            if condition_1 and condition_2:
                return 1
            else:
                return 0
        elif condition_1 or condition_2:
            return 1
        else:
            return 0

    iqei121['score1'] = iqei121.apply(calculate_score1, axis=1)
    iqei121['score2'] = iqei121.apply(calculate_score2, axis=1)
    iqei121['iqei1.2.1'] = np.where((iqei121['score1'] == 1) | (iqei121['score2'] == 1), 'Yes', 'No')
    
    if breakdown == None:
        bs.chi2(iqei121, 'iqei1.2.1', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
        
    order= ['Yes','No']
    xlabel = ' '
    if breakdown == None:
        iqei121_df = count_df(iqei121, 'iqei1.2.1', file_path, custom_order=order, index_name='IQEI1.2.1')
        title1 = 'IQEI1.2.1: % of parents who enrol their children in pre-primary \nand primary school at the right age regardless of sex and abilities'
        plot_bar(iqei121_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_IQEI1.2.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'iqei1.2.1'
        var_name = "IQEI1.2.1: % of parents who enrol their children in pre-primary \nand primary school at the right age regardless of sex and abilities"
        iqei121_df = table_breakdown(iqei121, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def proo211(pp_df, file_path, breakdown = None):
    proo211 = pp_df.copy(deep=True)

    def PROO211_score(row):
        score_mapping1 = {
        'Always happens': 3,
        'Often': 2,
        'Sometimes': 1,
        'Almost never': 0, 'It never happens':0
        }
        score_mapping2 = {
        'Strongly agree': 3,
        'Somewhat agree': 2,
        'Neither agree nor disagree': 1,
        'Somewhat disagree': 0, 'Strongly Disagree':0
        }
        score = 0
        columns1 = ['23-1','23-2','23-3','23-4','23-5','23-6','23-7','23-8','23-9','23-10','23-11','23-12','23-13','23-14','23-15','23-16',
                     '23-17']
        columns2 = ['24-1','24-2','24-3']
        non_na_count = 0  
    
        for col in columns1:
            if pd.notna(row[col]):
                score += score_mapping1.get(row[col], 0)
        for col in columns2:
            if pd.notna(row[col]):
                score += score_mapping2.get(row[col], 0)     
        return score / (len(columns1) + len(columns2))
    
    def label_score(score):
        if score <= 1:
            return 'Not positive parenting'
        elif 1 < score <= 2:
            return 'Somewhat positive parenting'
        else:
            return 'Positive parenting'

    proo211['average_score'] = proo211.apply(PROO211_score, axis=1)
    proo211['proo2.1.1'] = ""
    proo211['proo2.1.1'] = proo211['average_score'].apply(label_score)
    proo211.drop(columns=['average_score'], inplace=True)
    
    if breakdown == None:
        bs.chi2(proo211, 'proo2.1.1', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)    
    
    order= ['Not positive parenting','Somewhat positive parenting', 'Positive parenting']
    xlabel = ""
    if breakdown == None:
        proo211_df = count_df(proo211, 'proo2.1.1', file_path, custom_order=order, index_name='PROO2.1.1')
        title1 = 'PROO2.1.1 % of parents and caregivers who report using positive parenting practices with children'
        plot_bar(proo211_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO2.1.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'proo2.1.1'
        var_name = "PROO2.1.1 % of parents and caregivers who report using positive parenting practices with children"
        proo211_df = table_breakdown(proo211, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def proo225(pp_df, file_path, breakdown = None):
    proo225 = pp_df.copy(deep=True)

    def assign_pass_fail(row):
        selected_count = sum([row['27-1'], row['27-2'], row['27-3'], row['27-4'], row['27-o']])
        if (row['26'] == 'Yes') and (selected_count >= 2):
            return 'Know'
        else:
            return 'Do not know'
    proo225['proo2.2.5'] = ""
    proo225['proo2.2.5'] = proo225.apply(assign_pass_fail, axis=1)
    if breakdown == None:
        bs.chi2(proo225, 'proo2.2.5', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)    

    order= ['Know','Do not know']
    xlabel = ""
    if breakdown == None:
        proo225_df = count_df(proo225, 'proo2.2.5', file_path, custom_order=order, index_name='PROO2.2.5')
        title1 = 'PROO2.2.5: % of parents and caregivers who report that \nthey know the major protection risks their children face including in a crisis-affected context'
        plot_bar(proo225_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO2.2.5.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'proo2.2.5'
        var_name = "PROO2.2.5: % of parents and caregivers who report that they know the major protection risks their children face including in a crisis-affected context"
        proo225_df = table_breakdown(proo225, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
  
def ecdo253_parents(pp_df, file_path, breakdown = None):
    ecdo253_1 = pp_df.copy() 
    def assign_pass_fail(row):
        selected_count = sum([row['29-1'], row['29-2'], row['29-3'], row['29-4'], row['29-5'], row['29-6']])
        if (row['28'] == 'Yes'):
            if (row['29-7'] == 1):
                return 'Know'
            elif (selected_count >= 3):
                return 'Know'
            else: return 'Do not know'
        else:
            return 'Do not know'
    ecdo253_1['ecdo2.5.3'] = ""
    ecdo253_1['ecdo2.5.3'] = ecdo253_1.apply(assign_pass_fail, axis=1)
    
    if breakdown == None:
        bs.chi2(ecdo253_1, 'ecdo2.5.3', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05) 

    order= ['Know','Do not know']
    xlabel = ' '
    if breakdown == None:
        ecdo253_1_df = count_df(ecdo253_1, 'ecdo2.5.3', file_path, custom_order=order, index_name='ECDO2.5.3_1')
        title1 = '<Parents and Caregivers>\nECDO2.5.3: % of people who know what are good Infant and Young Child Feeding practices'
        plot_bar(ecdo253_1_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_ECDO2.5.3_1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'ecdo2.5.3'
        var_name = "<Parents and Caregivers>\nECDO2.5.3: % of people who know what are good Infant and Young Child Feeding practices"
        ecdo253_1_df = table_breakdown(ecdo253_1, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
 
def ecdo253_social(sw_df, file_path, breakdown = None):
    ecdo253_2 = sw_df.copy() 

    def assign_pass_fail(row):
        selected_count = sum([row['15-1'], row['15-2'], row['15-3'], row['15-4'], row['15-5'], row['15-6']])
        if (row['14'] == 'Yes'):
            if (selected_count >= 3):
                return 'Know'
            else: return 'Do not know'
        else:
            return 'Do not know'
    ecdo253_2['ecdo2.5.3'] = ""
    ecdo253_2['ecdo2.5.3'] = ecdo253_2.apply(assign_pass_fail, axis=1)
    if breakdown == None:
        bs.chi2(ecdo253_2, 'ecdo2.5.3', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05) 

    order= ['Know','Do not know']
    xlabel = ""
    if breakdown == None:
        ecdo253_2_df = count_df(ecdo253_2, 'ecdo2.5.3', file_path, custom_order=order, index_name='ECDO2.5.3_2')
        title1 = '<Social Workers>\nECDO2.5.3: % of people who know what are good Infant and Young Child Feeding practices'
        plot_bar(ecdo253_2_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_ECDO2.5.3_2.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'ecdo2.5.3'
        var_name = "<Social Workers>\nECDO2.5.3: % of people who know what are good Infant and Young Child Feeding practices"
        ecdo253_2_df = table_breakdown(ecdo253_2, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def ecdo259(pp_df, file_path, breakdown = None):
    ecdo259 = pp_df.copy()

    def assign_pass_fail(row):
        selected_count = sum([row[col] == 'Yes' for col in ['31-1', '31-2', '31-3', '31-4', '31-5', '31-6', '31-7', '31-8', '31-9', '31-10', '31-11']])
        if (row['30'] == 'Yes'):
            if (selected_count == 0):
                return 'Exclusively Breastfeeding'
            else: return 'Not Exclusively'
        else:
            return 'Not Exclusively'
    ecdo259['ecdo2.5.9'] = ""
    ecdo259['ecdo2.5.9'] = ecdo259.apply(assign_pass_fail, axis=1)
    if breakdown == None:
        bs.chi2(ecdo259, 'ecdo2.5.9', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05) 

    order= ['Exclusively Breastfeeding','Not Exclusively']
    xlabel = ""
    if breakdown == None:
        ecdo259_df = count_df(ecdo259, 'ecdo2.5.9', file_path, custom_order=order, index_name='ECDO2.5.9')
        title1 = 'ECDO2.5.9: % of parents/caregivers who report that their infants aged \nunder 6 months were fed exclusively with breast milk in the past 24 hours'
        plot_bar(ecdo259_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_ECDO2.5.9.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'ecdo2.5.9'
        var_name = "ECDO2.5.9: % of parents/caregivers who report that their infants aged under 6 months were fed exclusively with breast milk in the past 24 hours"
        ecdo259_df = table_breakdown(ecdo259, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def ecdo2510(pp_df, file_path, breakdown = None):
    ecdo2510 = pp_df[(pp_df['15']!='No') & (pp_df['15']!='Prefer not to answer')]

    def calculate_score(row):
        condition = ((row['33-I'] == 'Yes') | (row['33-II'] == 'Yes') | (row['33-III'] == 'Yes'))
        if condition:
            return 1
        else:
            return 0

    ecdo2510['score'] = ecdo2510.apply(calculate_score, axis=1)
    ecdo2510['ecdo2.5.10'] = np.where((ecdo2510['score'] == 1), 'Received', 'Not received')
    
    if breakdown == None:
        bs.chi2(ecdo2510, 'ecdo2.5.10', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05) 

    order= ['Received','Not received']
    xlabel = ""
    if breakdown == None:
        ecdo2510_df = count_df(ecdo2510, 'ecdo2.5.10', file_path, custom_order=order, index_name='ECDO2.5.10')
        title1 = 'ECDO2.5.10: % of infants aged 6 to 8 months of age who received \nsolid, semi-solid or soft foods during the previous day or night'
        plot_bar(ecdo2510_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_ECDO2.5.10.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'ecdo2.5.10'
        var_name = "ECDO2.5.10: % of infants aged 6 to 8 months of age who received \nsolid, semi-solid or soft foods during the previous day or night"
        ecdo2510_df = table_breakdown(ecdo2510, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def ecdo2512(pp_df, file_path, breakdown = None):
    ecdo2512 = pp_df.copy(deep=True)

    def calculate_score(row):
        condition_1 = any(row[col] == 1 for col in ['35-I-1', '35-I-2', '35-I-3', '35-I-4']) and row['35-I-5'] != 1
        condition_2 = any(row[col] == 1 for col in ['35-II-1', '35-II-2', '35-II-3', '35-II-4']) and row['35-II-5'] != 1
        if condition_1 and condition_2:
            return 1
        else:
            return 0

    ecdo2512['score'] = ecdo2512.apply(calculate_score, axis=1)
    ecdo2512['ecdo2.5.12'] = np.where((ecdo2512['score'] == 1), 'Have basic facilities', 'Do not have')
    if breakdown == None:
        bs.chi2(ecdo2512, 'ecdo2.5.12', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05) 

    order= ['Have basic facilities','Do not have']
    xlabel = ""
    if breakdown == None:
        ecdo2512_df = count_df(ecdo2512, 'ecdo2.5.12', file_path, custom_order=order, index_name='ECDO2.5.12')
        title1 = 'ECDO2.5.12: % of households, that are verified as using at least basic sanitation and hygiene facilities'
        plot_bar(ecdo2512_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_ECDO2.5.12.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'ecdo2.5.12'
        var_name = "ECDO2.5.12: % of households, that are verified as using at least basic sanitation and hygiene facilities"
        ecdo2512_df = table_breakdown(ecdo2512, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def ecdo2514(pp_df, file_path, breakdown = None):
    ecdo2514 = pp_df.copy(deep=True)

    def calculate_score(row):
        condition_36 = (any(row[f'36-{i}'] == 1 for i in range(1, 5)) and 
                    all(row[f'36-{i}'] != 1 for i in range(6, 10)))
    
        condition_37 = (any(row[f'37-{i}'] == 1 for i in range(1, 5)) and 
                    all(row[f'37-{i}'] != 1 for i in range(6, 10)))
    
        if condition_36 or condition_37:
            return 1
        else:
            return 0

    ecdo2514['score'] = ecdo2514.apply(calculate_score, axis=1)
    ecdo2514['ecdo2.5.14'] = np.where((ecdo2514['score'] == 1), 'Have appropriate water source', 'Do not have')
    if breakdown == None:
        bs.chi2(ecdo2514, 'ecdo2.5.14', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05) 

    order= ['Have appropriate water source','Do not have']
    xlabel = ' '
    if breakdown == None:
        ecdo2514_df = count_df(ecdo2514, 'ecdo2.5.14', file_path, custom_order=order, index_name='ECDO2.5.14')
        title1 = 'ECDO2.5.14: % of households where people drink water from a protected or treated water source'
        plot_bar(ecdo2514_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_ECDO2.5.14.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'ecdo2.5.14'
        var_name = "ECDO2.5.14: % of households where people drink water from a protected or treated water source"
        ecdo2514_df = table_breakdown(ecdo2514, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def ecdo271(pp_df, file_path, breakdown = None):
    ecdo271 = pp_df[pp_df['15']!='No']

    def calculate_score(row):
        condition_1 = (row['17-I'] == 'Yes' and 3 <= row['16-I'] <= 5)
        condition_2 = (row['17-II'] == 'Yes' and 3 <= row['16-II'] <= 5)
        if condition_1 or condition_2:
            return 1
        else:
            return 0

    ecdo271['score'] = ecdo271.apply(calculate_score, axis=1)
    ecdo271['ecdo2.7.1'] = np.where((ecdo271['score'] == 1), 'Enrolled', 'Not enrolled')
    if breakdown == None:
        bs.chi2(ecdo271, 'ecdo2.7.1', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05) 

    order= ['Enrolled','Not enrolled']
    xlabel = ' '
    if breakdown == None:
        ecdo271_df = count_df(ecdo271, 'ecdo2.7.1', file_path, custom_order=order, index_name='ECDO2.7.1')
        title1 = 'ECDO2.7.1: % of children (reached by the project) who have been enrolled in\n an Early Learning Programme (formal or non-formal)'
        plot_bar(ecdo271_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_ECDO2.7.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'ecdo2.7.1'
        var_name = "ECDO2.7.1: % of children (reached by the project) who have been enrolled in an Early Learning Programme (formal or non-formal)"
        ecdo271_df = table_breakdown(ecdo271, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def proo511_separated(sw_df, file_path, breakdown = None):
    proo511 = sw_df.copy(deep=True)

    score_map = {'Very confident performing task- I know what is expected of me and feel competent doing it' : 2,
                    'Somewhat confident performing task- I may need some support':1, "Not confident performing task-I will need assistance" : 0}

    columns_to_transform = [
    '13-1a', '13-1b', '13-1c', '13-2a', '13-2b', '13-2c', 
    '13-3a', '13-3b', '13-4a', '13-4b', '13-4c', '13-5a', 
    '13-5b', '13-5c', '13-6a', '13-6b', '13-6c', '13-7a', 
    '13-7b', '13-7c']
    
    for col in columns_to_transform:
        proo511[col] = proo511[col].map(score_map)

    columns_to_check1 = ['13-1a','13-1b','13-1c']
    proo511['1'] = proo511[columns_to_check1].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check2 = ['13-2a','13-2b','13-2c']
    proo511['2'] = proo511[columns_to_check2].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check3 = ['13-3a','13-3b']
    proo511['3'] = proo511[columns_to_check3].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check4 = ['13-4a','13-4b','13-4c']
    proo511['4'] = proo511[columns_to_check4].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check5 = ['13-4a','13-5b','13-5c']
    proo511['5'] = proo511[columns_to_check5].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check6 = ['13-6a','13-6b','13-6c']
    proo511['6'] = proo511[columns_to_check6].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check7 = ['13-7a','13-7b','13-7c']
    proo511['7'] = proo511[columns_to_check7].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    order= ['Confident','Not confident']
    xlabel = ' '
    if breakdown == None:
        proo511_df1 = count_df(proo511, '1', file_path, custom_order=order, index_name='PROO5.1.1-1')
        proo511_df2 = count_df(proo511, '2', file_path, custom_order=order, index_name='SRHO1.1.1-2')
        proo511_df3 = count_df(proo511, '3', file_path, custom_order=order, index_name='SRHO1.1.1-3')
        proo511_df4 = count_df(proo511, '4', file_path, custom_order=order, index_name='SRHO1.1.1-4')
        proo511_df5 = count_df(proo511, '5', file_path, custom_order=order, index_name='SRHO1.1.1-5')
        proo511_df6 = count_df(proo511, '6', file_path, custom_order=order, index_name='SRHO1.1.1-6')
        proo511_df7 = count_df(proo511, '7', file_path, custom_order=order, index_name='SRHO1.1.1-7')

        title1 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Violence against children>'
        title2 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<initial assessment>'
        title3 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Intimate partner violence>'
        title4 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Multi-sectorial services>'
        title5 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Follow up and monitoring>'
        title6 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Case management>'
        title7= 'PROO5.1.1 % of social service workers who are confident in their ability \n<Reporting>'

        plot_bar(proo511_df1, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO5.1.1-1.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df2, 'Percentage', title = title2, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO5.1.1-2.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df3, 'Percentage', title = title3, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO5.1.1-3.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df4, 'Percentage', title = title4, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO5.1.1-4.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df5, 'Percentage', title = title5, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO5.1.1-5.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df6, 'Percentage', title = title6, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO5.1.1-6.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df7, 'Percentage', title = title7, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO5.1.1-7.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        proo511_df1 = count_df(proo511, '1', file_path, custom_order=order, index_name=f'PROO5.1.1-1_{breakdown}')
        proo511_df2 = count_df(proo511, '2', file_path, custom_order=order, index_name=f'SRHO1.1.1-2_{breakdown}')
        proo511_df3 = count_df(proo511, '3', file_path, custom_order=order, index_name=f'SRHO1.1.1-3_{breakdown}')
        proo511_df4 = count_df(proo511, '4', file_path, custom_order=order, index_name=f'SRHO1.1.1-4_{breakdown}')
        proo511_df5 = count_df(proo511, '5', file_path, custom_order=order, index_name=f'SRHO1.1.1-5_{breakdown}')
        proo511_df6 = count_df(proo511, '6', file_path, custom_order=order, index_name=f'SRHO1.1.1-6_{breakdown}')
        proo511_df7 = count_df(proo511, '7', file_path, custom_order=order, index_name=f'SRHO1.1.1-7_{breakdown}')

        title1 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Violence against children>'
        title2 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<initial assessment>'
        title3 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Intimate partner violence>'
        title4 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Multi-sectorial services>'
        title5 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Follow up and monitoring>'
        title6 = 'PROO5.1.1 % of social service workers who are confident in their ability \n<Case management>'
        title7= 'PROO5.1.1 % of social service workers who are confident in their ability \n<Reporting>'

        plot_bar(proo511_df1, 'Percentage', title = title1, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_PROO5.1.1-1.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df2, 'Percentage', title = title2, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_PROO5.1.1-2.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df3, 'Percentage', title = title3, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_PROO5.1.1-3.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df4, 'Percentage', title = title4, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_PROO5.1.1-4.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df5, 'Percentage', title = title5, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_PROO5.1.1-5.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df6, 'Percentage', title = title6, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_PROO5.1.1-6.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(proo511_df7, 'Percentage', title = title7, xlabel = xlabel, output_file = f'Visuals/Plots/{breakdown}_PROO5.1.1-7.png', figsize=(18, 10), rotation=0, fontsize = 14)
    clear_output()        

def proo511(sw_df, file_path, breakdown = None):
    proo511 = sw_df.copy(deep=True)

    score_map = {'Very confident performing task- I know what is expected of me and feel competent doing it' : 2,
                    'Somewhat confident performing task- I may need some support':1, "Not confident performing task-I will need assistance" : 0}

    columns_to_transform = [
    '13-1a', '13-1b', '13-1c', '13-2a', '13-2b', '13-2c', 
    '13-3a', '13-3b', '13-4a', '13-4b', '13-4c', '13-5a', 
    '13-5b', '13-5c', '13-6a', '13-6b', '13-6c', '13-7a', 
    '13-7b', '13-7c']
    
    for col in columns_to_transform:
        proo511[col] = proo511[col].map(score_map)

    columns_to_check1 = ['13-1a','13-1b','13-1c']
    proo511['aa1'] = proo511[columns_to_check1].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check2 = ['13-2a','13-2b','13-2c']
    proo511['aa2'] = proo511[columns_to_check2].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check3 = ['13-3a','13-3b']
    proo511['aa3'] = proo511[columns_to_check3].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check4 = ['13-4a','13-4b','13-4c']
    proo511['aa4'] = proo511[columns_to_check4].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check5 = ['13-4a','13-5b','13-5c']
    proo511['aa5'] = proo511[columns_to_check5].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check6 = ['13-6a','13-6b','13-6c']
    proo511['aa6'] = proo511[columns_to_check6].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')
    columns_to_check7 = ['13-7a','13-7b','13-7c']
    proo511['aa7'] = proo511[columns_to_check7].mean(axis=1).apply(lambda x: 'Confident' if x >= 1 else 'Not confident')

    proo511['proo511_dc'] = 0    
    result = proo511[proo511['aa1'] == 'Confident']
    result = result.merge(proo511[proo511['aa2'] == 'Confident'], how='inner')
    result = result.merge(proo511[proo511['aa3'] == 'Confident'], how='inner')
    result = result.merge(proo511[proo511['aa4'] == 'Confident'], how='inner')
    result = result.merge(proo511[proo511['aa5'] == 'Confident'], how='inner')
    result = result.merge(proo511[proo511['aa6'] == 'Confident'], how='inner')
    result = result.merge(proo511[proo511['aa7'] == 'Confident'], how='inner')
    
    columns = ['aa1', 'aa2', 'aa3', 'aa4', 'aa5', 'aa6', 'aa7']
    indices_confident = [set(proo511[proo511[col] == 'Confident'].index) for col in columns]
    common_indices = set.intersection(*indices_confident)
    proo511.loc[list(common_indices), 'proo511_dc'] = 1
    
    if breakdown == None:
        bs.chi2(proo511, 'proo511_dc', '3', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)

    len_result = len(proo511)
    len_pass = len(result)
    len_fail = len_result - len_pass
    count_pass = len_pass
    count_fail = len_fail
    percent_pass = (count_pass / len_result) * 100
    percent_fail = (count_fail / len_result) * 100

    result_proo511 = pd.DataFrame({
    'Count': [count_pass, count_fail],
    'Percent': [percent_pass, percent_fail]
    }, index=['Confident', 'Not confident'])
    
    xlabel = ' '
    if breakdown == None:
        print(f'{percent_pass:.2f} percent of people meet the criteria for indicator SRHO1.1.1')
        title1 = 'PROO5.1.1 % of social service workers who are confident in their ability to manage and respond \nto cases of violence against children (VAC) and intimate partner violence (IPV)'
        plot_bar(result_proo511, 'Percent', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_PROO5.1.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'proo511_dc'
        var_name = "PROO5.1.1 % of social service workers who are confident in their ability to manage and respond to cases of violence against children (VAC) and intimate partner violence (IPV)"
        proo511_df = table_breakdown(proo511, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = None)
        
def leao112(yo_df, file_path, breakdown = None):
    leao112 = yo_df.copy()
    score_map = {'Yes' : 2,'Somewhat':1, "No" : 0}
    columns_to_transform = ['17','18','19','20','21','22','23','24','25','26','27','28','29','30']
    for col in columns_to_transform:
        leao112[col] = leao112[col].map(score_map)

    columns_to_check1 = ['17','18','19','20','21','22','23','24','25']
    columns_to_check2 = ['26','27','28','29','30']
    existing_columns_to_check1 = [col for col in columns_to_check1 if col in leao112.columns]
    leao112['score1'] = leao112[existing_columns_to_check1].sum(axis=1).apply(lambda x: 'Sufficient' if x >= 10 else 'Insufficient')
    existing_columns_to_check2 = [col for col in columns_to_check2 if col in leao112.columns]
    leao112['score2'] = leao112[existing_columns_to_check2].sum(axis=1).apply(lambda x: 'Sufficient' if x >= 6 else 'Insufficient')
    order= ['Sufficient','Insufficient']
    xlabel = ' '
    if breakdown == None:
        leao112_df1 = count_df(leao112, 'score1', file_path, custom_order=order, index_name='LEAO1.1.2-Situational Analysis')
        leao112_df2 = count_df(leao112, 'score2', file_path, custom_order=order, index_name='LEAO1.1.2-Advocacy Plan')
        title1 = 'LEAO1.1.2: % of youth organisations which demonstrate a critical understanding of \ntheir socio-cultural, economic political and legal context <Situational Analysis Assessment>'
        title2 = 'LEAO1.1.2: % of youth organisations which demonstrate a critical understanding of \ntheir socio-cultural, economic political and legal context <Advocacy Plan Assessment>'
        plot_bar(leao112_df1, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_LEAO1.1.2-1.png', figsize=(18, 10), rotation=0, fontsize = 14)
        plot_bar(leao112_df2, 'Percentage', title = title2, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_LEAO1.1.2-2.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'score1'
        var_name = "LEAO1.1.2: % of youth organisations which demonstrate a critical understanding of their socio-cultural, economic political and legal context <Situational Analysis Assessment>"
        leao112_df1 = table_breakdown(leao112, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        var = 'score2'
        var_name = "LEAO1.1.2: % of youth organisations which demonstrate a critical understanding of their socio-cultural, economic political and legal context <Advocacy Plan Assessment>"
        leao112_df2 = table_breakdown(leao112, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)        
        
def leao132(yo_df, file_path, breakdown = None):
    leao132_count = yo_df.copy()
    leao132_count_df1 = count_df(leao132_count, '33', file_path, index_name='LEAO1.3.2-Count')
    leao132_value = yo_df[yo_df['33']=='Yes']
    leao132_value['38-b']

    score_map = {'Yes' : 2,'Somewhat':1, "No" : 0}

    columns = ['38-b','38-c','39-a','39-b','40-a','40-b','41-a','41-b','42-a','42-b']
    for col in columns:
        leao132_value[col] = leao132_value[col].map(score_map)
    leao132_value['sum'] = leao132_value[columns].sum(axis=1)

    leao132_value['label'] = np.where(leao132_value['sum'] > 10, 'High quality actions', np.where((leao132_value['sum'] >= 7) & (leao132_value['sum'] <= 10), 'Acceptable','Low'))
    leao132_value.drop('sum', axis=1, inplace=True)
    order= ['High quality actions','Acceptable']  # Future: Low
    xlabel = ' '
    if breakdown == None:
        leao132_value = count_df(leao132_value, 'label', file_path, custom_order=order, index_name='LEAO1.3.2-Value')
        title1 = 'LEAO1.3.2: # and quality of young people’s organisations own and joint actions'
        plot_bar(leao132_value, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_LEAO1.3.2.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'label'
        var_name = "LEAO1.3.2: # and quality of young people’s organisations own and joint actions"
        leao132_value = table_breakdown(leao132_value, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def leao121(yo_df, file_path,  breakdown = None):
    score_map_43 = {'My organisation has a solid number of active members and new members of both sexes join the organization. Older/ more experienced members become facilitators and mentors for younger members; provision made for the inclusion and participation of different age groups of girls and boys.'
             : 3,'New members (girls and boys), particularly younger children (under 13 years) are regularly encouraged to join and play an active role in my organisation; children with disabilities and other excluded or vulnerable groups are active members'
             :2, "My organisation has few active members; many of the original members have left the organisation; no new members have joined in the past year" : 1,
            "My organisation has a small group of active members; a few new members have joined in the past year":0}
    score_map_44 = {'The members of my organisation meet on a regular basis and all members are clear about when and where we will meet. When need arises, ad hoc meetings are also organized'
             : 3,'My organisation provides regularly space for all members to meet and most members are informed about the meeting modalities (date, time, venue)'
             :2, "Our members hold meetings periodic meetings, but sometimes there are gaps. It happens that some of our members are not informed about when and where our meetings are held" 
                : 1,
            "My organisation has infrequent and/or very irregular meetings and there are frequent changes -there is no/little continuity.":0}
    score_map_45 = {'All girls and boys pro- actively share their views and play an active role in decision-making; all members have an equal voice irrespective of gender, age, disability, ethnicity, background; members decide about and share roles/ responsibilities flexibly, all members have a fair chance to represent their organisation in meetings, workshops and policy events; processes for fair election and representation are established; members understand decisions and how they are developed'
             : 3,'Most members actively participate in decision-making; many different girls and boys have a chance to represent their organisation in meetings, workshops and training; all members are encouraged to have representation and facilitation skills; there is a fair election system; decisions are shared with all members'
             :2, "Decisions are made by some members; in most situations the same few girls or boys represent the organisation in meetings, workshops and training; there is an election system for group leaders, but it does not give an equal chance to all members" 
                : 1,
            "Decisions are mostly made by a small group (2-5) of members; no fair system of election for representation is developed; the few 'same' members tend to represent the organisation in meetings, workshops and training":0}
    
    score_map_46 = {"Clear channels of communication and information sharing exist amongst all members; all members have access to comprehensive information concerning their organisation, and issues affecting them; open, respectful communication amongst all members; regular documentation of activities and experiences; clear system for keeping meeting minutes, reports and financial records accessible to all members"
             : 3,"All members share information and communicate well; members have access to information on some issues affecting them, but not all; There is regular documentation activities and experiences"
             :2, "Most members have little access to information on key issues affecting them; there is no clear communication and information sharing within the organisation and no documentation of activities" 
                : 1,
            "Members have a system of information sharing, but sometimes communication breakdowns occur and members have access to some kinds of information, but not enough; there is occasional documentation of activities, but not regularly":0}
    score_map_47 = {"Members regularly identify issues concerning them; they analyse the root causes and impact of problems/ issues affecting themThey focus on key priority issues and develop and implement realistic action plans to achieve positive results"
             : 3,"Members are involved in analysis of issues and understand the root causes and impact of issues affecting them; they make action plans to solve the problems - most of these plans are implemented, but some are not"
             :2, "Members’ meeting is identified as 'talking shop' where issues are discussed, but rarely followed up by action; only few issues are followed up by action" 
                : 1,
            "Very little analysis of issues raised by members and infrequent action taken; between meetings members undertake only minimal activities or action; some of the young people do not feel any benefits resulting from their group":0}
    score_map_48 = {"All members have a choice regarding the nature and degree of their participation; the barriers adolescent girls/ young women face in their participation are addressed and they are able to meaningfully participate along-side the adolescent boys and young men in my organisation, we regularly assess the risks involved in our activities and develop strategies to mitigate them"
             : 3,"Most of our members are clear about their participation and have choices; in some situations, young people assess risks and make sure they are protected; adolescent boys/ young men and adolescent girls/ young women are equally active in my organisation"
             :2, "Boys and girls have choices about their participation (when, what, how, where), but adolescent boys and young men are more active than the adolescent girls/ young women in my organisation.  Our members are sometimes aware of risks associated with participation and sometimes make informed choices" 
                : 1,
            "Members are not clear about the purpose of their participation; they may be placing themselves at risk as a result of their participation; adolescent girls and young women hardly participate in my organisation":0}
    score_map_49 = {"Members are confident in developing activities which ensure an equal participation of all girls and boys of different ages and abilities"
             : 3,"There are strong efforts to develop activities in such a way that girls and boys of different ages and abilities can participate; e.g. by using creative activities like drama or drawing"
             :2, "We occasionally consider ways of how to involve girls and boys of different ages and abilities in our activities, but it is mainly a small group of members who decide on the activities" 
                : 1,
            "In my organisation, we usually organise our activities in such a way that it makes it hard for younger members, those with disabilities, and/or who are illiterate to participate":0}
    score_map_50 = {"The young people’s organisation has access to resources, training, and support which will enable it to continue for a long time; the members have mobilised support and resources from their local community; the members have their own fundraising activities and manage their own funds as instructed by an external agency"
             : 3,"The organisation has access to information, resources, training and support; funds are mobilised from local community sources; members have their own fundraising activities. The organisation manages its own funds with occasional support from an external organisation/ institution"
             :2, "The organisation has some resources, but not enough; The members have mobilised some resources from the local community, funds are managed by an external agency with inputs/ justifications from the youth organisations" 
                : 1,
            "The organisation is constrained by its lack of resources and is totally dependent on resources from an external (national, international) agency; the members have been unable to mobilise support, materials or space from their local community; financial resources are not managed by the youth organisation, but by an external agency":0}
    score_map_51 = {"My organisation is regularly addressing issues around gender equality, all members have been trained and have the necessary knowledge on gender equality and are committed to its advancement. Girls and boys participate equally in the planning, organisation and implementation of actions and are self-organized and contribute according to their capacity. Our members are motivated to share responsibilities equally"
             : 3,"The members have awareness on gender equality related issues. They are taken seriously and discussed openly by all, but not all members have received training on gender topics. There is an equal presentation of girls and boys in organisational leadership positions, but more influential positions are hold by boys/ young men"
             :2, "Gender related discriminations such as the absence of girls in organisational leadership position are discussed, but have not (yet) been addressed. Leadership positions in the organizations are mostly or exclusively filled by boys/ young men, but there is recognition that there should be an equal presentation of girls and boys" 
                : 1,
            "Gender equality at organisational level has not been a priority area; most members believe that gender equality is not a critical issue that matters to their work and no one has been trained on gender equality in the past two years. Leadership position in the organisations are mostly or exclusively filled by boys/ young men":0}
    score_map_52 = {"Regular reflection on strengths and weaknesses of organisation and lessons applied to improve the organisation; documented evidence of how the organisation has learnt from experiences (both positive and negative); systems in place for documentation, monitoring and evaluation; boys and girls regularly gather views from other members, and adult duty bearers about the impact of their activities"
             : 3,"Members regularly reflect on their strengths, weaknesses and how to improve; The young people’s organisation has a system for monitoring progress of their action plans: members are willing to learn from their mistakes and the organisation has a system of documenting the progress of its activities"
             :2, "Members occasionally reflect on their strengths and weaknesses, but not often; members don't apply many of their learnings; there is not much documentation of activities, accomplishments and learning" 
                : 1,
            "There are no systems in place for group learning, reflection, monitoring or evaluation; members do not share or learn from mistakes; there is no after-action discussion and neither documentation nor evaluation of activities":0}
    score_map_53 = {"Members communicate respectfully with each other and give each other constructive feedback; different and critical ideas are valued and discussed; the organisation has a process for solving conflicts and disagreement between members which enables the identification of participatory solutions"
             : 3,"Critical and different ideas are discussed and reflected on.  in situations of conflicts and disagreements, all members can express their viewpoints and are treated fairly, there is, however, no process for solving conflicts"
             :2, "There is room to share different and critical ideas, but they are not taken into consideration in decision-making processes. Disagreement and conflicts are discussed, but not always solved to the satisfaction of all" 
                : 1,
            "There is little room for different and critical ideas and members are hesitant to share their viewpoints. When conflicts and disagreements appear; they are dealt with by an adult chaperon or remain undiscussed":0}
    
    leao121 = yo_df.copy()

    score_maps = [score_map_43, score_map_44, score_map_45, score_map_46, score_map_47,
             score_map_48, score_map_49, score_map_50, score_map_51, score_map_52, score_map_53]
    columns = ['43','44','45','46','47','48','49','50','51','52','53']
    for col, score_map in zip(columns, score_maps):
        leao121[col] = leao121[col].map(score_map)
    
    leao121['sum'] = leao121[columns].sum(axis=1)

    leao121['label'] = np.where(leao121['sum'] >= 18, 'High capacity', 'Low capacity')
    leao121.drop('sum', axis=1, inplace=True)
    order= ['High capacity', 'Low capacity']
    xlabel = ' '
    if breakdown == None:
        leao121 = count_df(leao121, 'label', file_path, custom_order=order, index_name='LEAO1.2.1')
        title1 = 'LEAO1.2.1: % of young people’s organisations whose organisational practices and \nstructures are gender transformative, inclusive and fit for purpose'
        plot_bar(leao121, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_LEAO1.2.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'label'
        var_name = "LEAO1.3.2: # and quality of young people’s organisations own and joint actions"
        leao121_df = table_breakdown(leao121, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)
        
def sdg1621(cay_df, file_path, breakdown = None):
    sdg1621 = cay_df[(cay_df['2'] >= 1) & (cay_df['2'] <= 17)]

    def sdg1621_score(row):
        score_mapping_1 = {
        'Yes': 1,
        'No': 0,
        'Unsure': 0,
        'Prefer not to answer': 0
        }
        score = 0
        columns1 = ['70','71','72', '73']
        for col in columns1:
            score += score_mapping_1.get(row[col], 0)
        if any(row[col] == 1 for col in columns1):
            score += 1
        return score

    def label_score(score):
        if score > 0:
            return 'Experienced'
        else:
            return 'No experience'
    sdg1621['score'] = sdg1621.apply(sdg1621_score, axis=1)
    sdg1621['sdg1621'] = sdg1621['score'].apply(label_score)
    
    if breakdown == None:
        bs.chi2(sdg1621, 'sdg1621', '4', change_var = {'Female':0, 'Male' : 1}, alpha = 0.05)
        
    sdg1621.drop(columns=['score'], inplace=True)
    order= ['Experienced','No experience']
    xlabel = ' '
    if breakdown == None:
        sdg1621_df = count_df(sdg1621, 'sdg1621', file_path, custom_order=order, index_name='SDG16.2.1')
        title1 = 'SDG 16.2.1: Percentage of children aged 1-17 years who experienced any physical \nand/or psychological aggression by caregivers in the past month'
        plot_bar(sdg1621_df, 'Percentage', title = title1, xlabel = xlabel, output_file = 'Visuals/Plots/oveall_SDG16.2.1.png', figsize=(18, 10), rotation=0, fontsize = 14)
    else:
        var = 'sdg1621'
        var_name = "SDG 16.2.1: Percentage of children aged 1-17 years who experienced any physical and/or psychological aggression by caregivers in the past month"
        sdg1621_df = table_breakdown(sdg1621, breakdown, var, sheet_name=var, var_name = var_name, file_path = file_path, var_order = order)       
        

        
        











        
 
    
 
    
 
        
        
        
        
        
        
        
        