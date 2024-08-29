#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Aug 25 14:23:14 2024

@author: ijeong-yeon
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from PIL import Image
from pandas.plotting import table
from IPython.display import clear_output
import warnings
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.tools.tools import add_constant
from scipy.stats import normaltest
from statsmodels.stats.diagnostic import lilliefors
import statsmodels.api as sm
from scipy import stats
from scipy.stats import f_oneway
from statsmodels.formula.api import ols

""" for statistics 
empty_df = pd.DataFrame()
file_path = 'Visuals/Tables/stats_sheet.xlsx'   
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    empty_df.to_excel(writer, sheet_name='basic', index=False)
"""

def vif_test(df, file_path):
    df = add_constant(df)

    vif_data = pd.DataFrame()
    vif_data["Variable"] = df.columns
    vif_data["VIF"] = [variance_inflation_factor(df.values, i) for i in range(df.shape[1])]
    
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        vif_data.to_excel(writer, sheet_name='VIF', index=True, header=True)
    wb = load_workbook(file_path)
    ws = wb['VIF']
    ws.insert_rows(1)
    ws['A1'] = 'VIF Test Result'
    ws['A1'].font = Font(bold=True)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)

    print(vif_data)
    
def normality_test(df, file_path, alpha=0.01):
    
    def hypothesis_statement(p_value, alpha):
        c = 100.0*alpha
        if p_value < alpha:
            return f"We reject the Null Hypothesis in favour of the Alternative Hypothesis at the {c}% level, p-value = {p_value}", False
        else:
            return f"We accept the Null Hypothesis at the {c}% level, p-value = {p_value}", True

    _1, p_value1 = normaltest(df.mscore)
    var1 = "D'Agostino and Pearson's K-squared"
    _2, p_value2 = lilliefors(df.mscore,dist="norm")
    var2 = 'Lilliefors'
    
    result_df = pd.DataFrame({
    'Test': [var1, var2],
    'Statistics': [_1, _2],
    'P-value': [p_value1, p_value2]})
    
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        result_df.to_excel(writer, sheet_name='Normality Test', index=True, header=True)
    wb = load_workbook(file_path)
    ws = wb['Normality Test']
    ws.insert_rows(1)
    ws['A1'] = 'Normality Test Result'
    ws['A1'].font = Font(bold=True)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)

    return result_df
    

def OLS(df, dependent_var, independent_vars):

    X = df[independent_vars]
    y = df[dependent_var]

    X = sm.add_constant(X)
    model = sm.OLS(y, X).fit()
    
    print(model.summary())
    
def chi2(df2, dependent_var, col, change_var, alpha=0.05):
    if change_var != None:
        df2[col] = df2[col].replace(change_var)
    contingency_table = pd.crosstab(df2[dependent_var], df2[col])
    chi2, p_value, _, _ = stats.chi2_contingency(contingency_table)
        
    print(f"Chi-square test statistic: {chi2}")
    print(f"P-value: {p_value}")
    var = 'Chi-square Test'
        
    if p_value < alpha:
        print(f"Variable: {col} | There is a significant association between {dependent_var} and {col}")
        print("")
    else:
        print(f"Variable: {col} | There is not a significant association between {dependent_var} and {col}")
        print("")
            

    result_df = pd.DataFrame({
        'Test': [var],
        'Statistics': [chi2],
        'P-value': [p_value]
    })
    
    return result_df    
    
    
def t_test(df2, dependent_var, col, change_var, alpha = 0.05, bootstrap = False):
    #  for example: change_var = {'Female':0, 'Male' : 1}
    df2[col] = df2[col].replace(change_var)
    df = df2.loc[:, [dependent_var, col]]
    unique_values = sorted(set(change_var.values()))
    index_0 = []
    index_1 = []

    for index in range(len(df)):
        if df.loc[index, col] == unique_values[0]:
            index_0.append(index)
        else: index_1.append(index)

    df_low = df.iloc[index_0, 0]
    df_high = df.iloc[index_1, 0]

    mean_low = np.mean(df_low)
    mean_high = np.mean(df_high)
    
    if bootstrap == True:
        n_bootstrap = 1000
        np.random.seed(101)

        group1 = np.array(df_low)
        group2 = np.array(df_high)

        size1 = len(group1) 
        data1 = np.array([np.mean(group1[np.random.randint(0,size1,size=size1)]) for _ in range(n_bootstrap)])

        size2 = len(group2)
        data2 = np.array([np.mean(group2[np.random.randint(0,size2,size=size2)]) for _ in range(n_bootstrap)])
    
        t_statistic, p_value = stats.ttest_ind(data1, data2)

        print(f"Mean value: {mean_high:.4f} vs {mean_low:.4f}")
        print("t-statistic:", t_statistic)
        print("p-value:", p_value)
        var = 'T-test with Bootstrap'
        
    else:
        t_statistic, p_value = stats.ttest_ind(df_low, df_high)

        print(f"Mean value: {mean_high:.4f} vs {mean_low:.4f}")
        print("t-statistic:", t_statistic)
        print("p-value:", p_value)
        var = 'T-test Result'

    if p_value < alpha:
        print(f"Variable : {col} | There is a significant statistical difference between {dependent_var} and {col}")
        print("")
    else:
        print(f"Variable : {col} | There is not a significant statistical difference between {dependent_var} and {col}")
        print("")
        
    result_df = pd.DataFrame({
    'Test': var,
    'Statistics': t_statistic,
    'P-value': p_value})
    """
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        result_df.to_excel(writer, sheet_name='T-test', index=True, header=True)
    wb = load_workbook(file_path)
    ws = wb['T-test']
    ws.insert_rows(1)
    ws['A1'] = 'T-test Result'
    ws['A1'].font = Font(bold=True)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
    """

    return result_df
    
def f_test(df2, dependent_var, col, change_var, file_path, alpha = 0.05, bootstrap = False):
    df2[col] = df2[col].replace(change_var)
    df = df2.loc[:, [dependent_var, col]]
    unique_values = sorted(set(change_var.values()))
    
    index_dict = {f'index_{val}': df.index[df[col] == val].tolist() for val in unique_values}
    

    for index, val in zip(range(len(df)), unique_values):
        if df.loc[index, col] == val:
            index_dict[f'index_{val}'].append(index)
            
    groups = [df.loc[indices, dependent_var] for indices in index_dict.values()]
    
    f_statistic, p_value = stats.f_oneway(*groups)
    
    print(f'F-statistic: {f_statistic}')
    print(f'P-value: {p_value}')
    
    if p_value < alpha:
        print("Reject the null hypothesis: There are significant differences between groups.")
    else:
        print("Fail to reject the null hypothesis: No significant differences between groups.")
        
    if bootstrap == True:
        n_bootstrap = 1000
        np.random.seed(101)
        
        bootstrap_f_stats = []
        for _ in range(n_bootstrap):
            bootstrap_samples = [np.random.choice(group, size=len(group), replace=True) for group in groups]
            bootstrap_f_stat, _ = stats.f_oneway(*bootstrap_samples)
            bootstrap_f_stats.append(bootstrap_f_stat)
        
        # Calculate bootstrap p-value
        bootstrap_f_stats = np.array(bootstrap_f_stats)
        bootstrap_p_value = np.mean(bootstrap_f_stats >= f_statistic)
        
        print(f'Bootstrap p-value: {bootstrap_p_value}')
        
        if bootstrap_p_value < alpha:
            print("Reject the null hypothesis: Significant differences between groups based on bootstrap.")
        else:
            print("Fail to reject the null hypothesis: No significant differences between groups based on bootstrap.")
        var = 'F-Test with bootstrap'
        result_df = pd.DataFrame({
     'Test': var,
     'Statistics': bootstrap_f_stats,
     'P-value': bootstrap_p_value})
    else:
        var = 'F-Test'
        result_df = pd.DataFrame({
     'Test': var,
     'Statistics': f_statistic,
     'P-value': p_value})
    
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        result_df.to_excel(writer, sheet_name='F-test', index=True, header=True)
    wb = load_workbook(file_path)
    ws = wb['F-test']
    ws.insert_rows(1)
    ws['A1'] = 'F-test Result'
    ws['A1'].font = Font(bold=True)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)

    return result_df
    

